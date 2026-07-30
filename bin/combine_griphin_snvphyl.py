#!/usr/bin/env python3

#disable cache usage in the Python so __pycache__ isn't formed. If you don't do this using 'nextflow run cdcgov/phoenix...' a second time will causes and error
import sys
sys.dont_write_bytecode = True
import pandas as pd
import argparse
import glob
import os
import re
import openpyxl
from openpyxl.styles import PatternFill, Font
from copy import copy
import numpy as np
from species_complexes import collapse_species_complex

# Function to get the script version
def get_version():
    return "1.0.0"

def parseArgs(args=None):
    parser = argparse.ArgumentParser(description="Combine griphin file with snvmatrix information.")
    parser.add_argument("-g", "--griphin", required=True, help="Input GRiPHin_Summary.xlsx file")
    parser.add_argument('-b', '--blind_list', default=None, required=False, dest='blind_list', help='CSV file with a list of sample_name,new_name. This option will output the new_name rather than the sample name to "blind" reports.')
    parser.add_argument('-w', '--window_size', default=None, required=False, dest='window_size', help='Window size for SNVPhyl analysis.')
    parser.add_argument('--combine_complex', default=False, action='store_true', required=False, dest='combine_complex', help='Group species belonging to the same species complex (e.g. Citrobacter freundii complex) into a single sheet named after the complex.')
    parser.add_argument('--version', action='version', version=get_version())# Add an argument to display the version
    return parser.parse_args()

def blind_map(control_file):
    """If you passed a file to -c this will swap out sample names to 'blind' the WGS_IDs in the final excel file."""
    rename_mapping = {}
    with open(control_file, 'r') as controls:
        header = next(controls) # skip the first line of the samplesheet
        for line in controls:
            old_sample_name = line.split(",")[0]
            new_sample_name = line.split(",")[1].rstrip("\n")
            if new_sample_name != '' and new_sample_name != old_sample_name:
                rename_mapping[old_sample_name] = new_sample_name
    return rename_mapping

def copy_sheet_dimensions(src_ws, dst_ws):
    # Column widths
    for col, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col].width = dim.width

def copy_header_merges(src_ws, dst_ws, max_row=1):
    """
    Copy merged cell ranges from src_ws to dst_ws but only if the merged range starts within the first `max_row` rows.
    """
    for merged_range in list(src_ws.merged_cells.ranges):
        if merged_range.min_row <= max_row:
            dst_ws.merge_cells(str(merged_range))

def sanitize_sheet_name(taxa_name):
    """Convert taxa name to valid Excel sheet name."""
    # Replace space with underscore for "genus species" format
    sheet_name = taxa_name.replace(" ", "_")
    # Remove any special characters that Excel doesn't allow in sheet names
    # Excel doesn't allow: / \ ? * [ ] :
    sheet_name = re.sub(r'[/\\?*\[\]:]', '', sheet_name)
    # If longer than 31 characters, truncate genus to first letter
    if len(sheet_name) > 31:
        parts = sheet_name.split("_")
        if len(parts) >= 2:
            sheet_name = parts[0][0] + "_" + "_".join(parts[1:])
        # If still too long, truncate to 31 characters
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
    return sheet_name

def copy_cell_style(source_cell, target_cell):
    target_cell.font = copy(source_cell.font)
    target_cell.border = copy(source_cell.border)
    target_cell.fill = copy(source_cell.fill)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)
    target_cell.alignment = copy(source_cell.alignment)

def remove_blank_columns(new_sheet, original_sheet, num_data_rows):
    """
    Remove columns from new_sheet that are entirely blank across all data rows (row 3 onward). Only columns AFTER the "No_AR_Genes_Found" column
    (found in row 2) are eligible for removal - columns at or before it are always kept regardless of blank status. Row-1 merged header ranges are
    remapped/shrunk to match the surviving columns. If a merge's entire column span is removed, an error is printed and that merge is dropped
    rather than re-added. After removal, columns after the boundary are auto-sized to fit their content. Returns nothing; modifies new_sheet in place.
    """
    max_col = original_sheet.max_column
    last_data_row = 2 + num_data_rows  # rows 1-2 are headers, data starts at row 3
    # Find the "No_AR_Genes_Found" column in row 2 - columns at or before this index are never eligible for removal.
    boundary_col_idx = None
    for col_idx in range(1, max_col + 1):
        if new_sheet.cell(row=2, column=col_idx).value == "No_AR_Genes_Found":
            boundary_col_idx = col_idx
            break
    if boundary_col_idx is None:
        print("WARNING: Could not find 'No_AR_Genes_Found' column in row 2. Skipping blank column removal.")
        return
    # Determine which columns (1-indexed) have at least one non-blank value among the data rows. Columns at or before boundary_col_idx are always kept; only columns after it are checked for blankness.
    cols_to_keep = []
    for col_idx in range(1, max_col + 1):
        if col_idx <= boundary_col_idx:
            cols_to_keep.append(col_idx)
            continue
        has_value = False
        for row_idx in range(3, last_data_row + 1):
            cell_value = new_sheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None and str(cell_value).strip() != '':
                has_value = True
                break
        if has_value:
            cols_to_keep.append(col_idx)
    if len(cols_to_keep) == max_col:
        # No columns removed, but still auto-size the columns after the boundary since their content may still warrant a resize.
        autosize_columns_after(new_sheet, boundary_col_idx, last_data_row)
        return
    # Capture existing row-1 merge ranges (old column indices) before we start deleting columns, so we can remap them afterward.
    old_merges = []
    for merged_range in list(new_sheet.merged_cells.ranges):
        if merged_range.min_row == 1 and merged_range.max_row == 1:
            old_merges.append((merged_range.min_col, merged_range.max_col, merged_range.min_row, merged_range.max_row))
            new_sheet.unmerge_cells(str(merged_range))
    # Build old-column-index -> new-column-index mapping (1-indexed) for surviving columns only.
    old_to_new_col = {old_idx: new_idx for new_idx, old_idx in enumerate(cols_to_keep, start=1)}
    # Delete blank columns, from rightmost to leftmost so indices don't shift out from under us mid-loop.
    cols_to_remove = [c for c in range(1, max_col + 1) if c not in old_to_new_col]
    for col_idx in sorted(cols_to_remove, reverse=True):
        new_sheet.delete_cols(col_idx)
    # Remap and re-apply row-1 merges against the new column positions.
    for min_col, max_col_old, min_row, max_row in old_merges:
        surviving_cols_in_range = sorted(
            old_to_new_col[c] for c in range(min_col, max_col_old + 1) if c in old_to_new_col
        )
        if not surviving_cols_in_range:
            print(f"ERROR: Merged header spanning columns {min_col}-{max_col_old} in row {min_row} had all underlying columns removed as blank. This should not happen.")
            continue
        new_min_col = surviving_cols_in_range[0]
        new_max_col = surviving_cols_in_range[-1]
        if new_min_col != new_max_col:
            new_sheet.merge_cells(start_row=min_row, start_column=new_min_col, end_row=max_row, end_column=new_max_col)
        # if new_min_col == new_max_col, only one column survived - no merge needed
    # Auto-size columns after the boundary to fit their content, since the original column widths no longer apply post-removal.
    new_boundary_col_idx = old_to_new_col[boundary_col_idx]
    autosize_columns_after(new_sheet, new_boundary_col_idx, last_data_row)


def autosize_columns_after(sheet, boundary_col_idx, last_data_row):
    """Set the width of every column after boundary_col_idx to fit the widest value found in that column, checking both header rows (1-2) and
    data rows (3 through last_data_row). Uses openpyxl's column letter addressing since column_dimensions is keyed by letter, not index.
    """
    max_col = sheet.max_column
    for col_idx in range(boundary_col_idx + 1, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, last_data_row + 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_len = max(max_len, len(str(cell_value)))
        if max_len > 0:
            # Add a little padding so text isn't flush against cell borders.
            sheet.column_dimensions[col_letter].width = max_len + 2

def create_taxa_sheets(workbook, combine_complex):
    """Create separate sheets for each unique taxa from Final_Taxa_ID column. 
    If combine_complex is True, species belonging to a recognized species complex (see species_complexes.py) are grouped together into a single sheet named after the complex.
    Columns that are entirely blank across a sheet's data rows are removed, and row-1 merged headers are shrunk/dropped to match.
    """
    original_sheet = workbook.active
    # Find the Final_Taxa_ID column
    taxa_col_idx = None
    for col_idx, cell in enumerate(original_sheet[2], start=1):  # Check row 2 for column names
        if cell.value == "Final_Taxa_ID":
            taxa_col_idx = col_idx
            break
    if taxa_col_idx is None:
        raise ValueError("Could not find 'Final_Taxa_ID' column in the Excel file")
    # Get unique taxa (starting from row 3, after 2 header rows)
    taxa_set = set()
    taxa_rows = {}  # Dictionary to store rows for each taxa (keyed by collapsed taxa name)
    for row_idx in range(3, original_sheet.max_row - 10):
        taxa_value = original_sheet.cell(row=row_idx, column=taxa_col_idx).value
        # Check for blank/NA taxa
        if taxa_value is None or str(taxa_value).strip() == '' or str(taxa_value).upper() == 'NA':
            raise ValueError(f"Row {row_idx} has blank/NA value in Final_Taxa_ID column. This should not happen.")
        # Collapse species complex members into their complex-level name so they land in the same sheet. Only done when --combine_complex is passed.
        collapsed_taxa = collapse_species_complex(taxa_value) if combine_complex else str(taxa_value).strip()
        taxa_set.add(collapsed_taxa)
        if collapsed_taxa not in taxa_rows:
            taxa_rows[collapsed_taxa] = []
        taxa_rows[collapsed_taxa].append(row_idx)
    # Create a sheet for each unique taxa (or complex)
    taxa_sheet_mapping = {}  # Map taxa to sheet names
    for taxa in sorted(taxa_set):
        sheet_name = sanitize_sheet_name(taxa)
        # Create new sheet
        new_sheet = workbook.create_sheet(title=sheet_name)
        # Copy sets dimensions BEFORE writing cells
        copy_sheet_dimensions(original_sheet, new_sheet)
        # Copy the 2 header rows
        for row_idx in range(1, 3):
            for col_idx in range(1, original_sheet.max_column + 1):
                source_cell = original_sheet.cell(row=row_idx, column=col_idx)
                target_cell = new_sheet.cell(row=row_idx, column=col_idx)
                # Copy value
                target_cell.value = source_cell.value
                # Copy formatting
                if source_cell.has_style:
                    copy_cell_style(source_cell, target_cell)
        # Copy sets header merges
        copy_header_merges(original_sheet, new_sheet)
        # Copy data rows for this taxa
        new_row_idx = 3
        for original_row_idx in taxa_rows[taxa]:
            for col_idx in range(1, original_sheet.max_column + 1):
                source_cell = original_sheet.cell(row=original_row_idx, column=col_idx)
                target_cell = new_sheet.cell(row=new_row_idx, column=col_idx)
                # Copy value
                target_cell.value = source_cell.value
                # Copy formatting
                if source_cell.has_style:
                    copy_cell_style(source_cell, target_cell)
            new_row_idx += 1
        # Remove columns that are entirely blank across this sheet's data rows, shrinking/dropping row-1 merged headers as needed.
        remove_blank_columns(new_sheet, original_sheet, len(taxa_rows[taxa]))
        # Store mapping of taxa to sheet name
        taxa_sheet_mapping[taxa.replace(" ", "_")] = sheet_name
    return taxa_sheet_mapping


def extract_taxa_from_filename(filename, combine_complex):
    """Extract taxa name from SNVMatrix filename.
    When combine_complex is True, expects complex-level files to include the literal "complex" segment, e.g. "Citrobacter_freundii_complex_ST169_snvMatrix.tsv" and "All_Citrobacter_freundii_complex_Isolates_snvMatrix.tsv".
    """
    basename = os.path.basename(filename)
    # Remove _snvMatrix.tsv
    taxa = basename.replace('_snvMatrix.tsv', '')
    # Remove All_ prefix if present
    if taxa.startswith('All_'):
        taxa = taxa[4:]  # Remove 'All_'
    # Remove _Isolates suffix if present
    taxa = taxa.replace('_Isolates', '')
    # Extract genus_species (or genus_species_complex) in case there are additional parts like ST307.
    # Complex names have an extra "complex" part (e.g. "Citrobacter_freundii_complex_ST169"), so take the first 3 underscore-parts for those (only when combine_complex is on) and the first 2 otherwise.
    parts = taxa.split('_')
    if combine_complex and len(parts) >= 3 and parts[2] == 'complex':
        taxa = f"{parts[0]}_{parts[1]}_{parts[2]}"
    elif len(parts) >= 2:
        taxa = f"{parts[0]}_{parts[1]}"
    return taxa


def append_tsv_to_excel(workbook, snvmatrices, result_dict, blind_list, taxa_sheet_mapping, window_size, snv_range, combine_complex):
    """Append SNVPhyl matrices to their corresponding taxa sheets."""
    # Group SNVMatrix files by taxa
    taxa_files = {}
    for snvmatrix in snvmatrices:
        taxa = extract_taxa_from_filename(snvmatrix, combine_complex)
        if taxa not in taxa_files:
            taxa_files[taxa] = []
        taxa_files[taxa].append(snvmatrix)
    # If blind_list exists, create mapping
    rename_mapping = None
    if blind_list is not None:
        rename_mapping = blind_map(blind_list)
    # Process each taxa's SNVMatrix files
    for taxa, files in taxa_files.items():
        # Find the corresponding sheet
        sheet_name = taxa_sheet_mapping.get(taxa)
        if sheet_name is None:
            print(f"WARNING: No matching taxa sheet found for SNVMatrix files with taxa '{taxa}'. Files: {files}")
            continue
        sheet = workbook[sheet_name]
        # Determine starting row
        start_row = sheet.max_row + 2
        # Define formatting
        bold_font = Font(bold=True)
        count = 0
        for snvmatrix in files:
            # Load the TSV file data
            snvmatrix_df = pd.read_csv(snvmatrix, sep='\t', dtype='str')
            if rename_mapping is not None:
                snvmatrix_df['WGS_ID'] = snvmatrix_df['WGS_ID'].replace(rename_mapping)
                snvmatrix_df = snvmatrix_df.rename(columns=rename_mapping)
            # Add section header on first file
            if count == 0:
                snvmatrix_cell = sheet.cell(row=start_row, column=1, value="SNVPhyl Analysis: SNV Matrices")
                snvmatrix_cell.fill = PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid")
                snvmatrix_cell.font = Font(bold=True)
                sheet.cell(row=start_row, column=2).fill = PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid")
                sheet.cell(row=start_row, column=3).fill = PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid")
                count = 1
            # Derive seq_type from filename
            seq_type = os.path.basename(snvmatrix).replace('_snvMatrix.tsv', '')
            # Write seq_type label
            sheet.cell(row=start_row + 1, column=1, value=seq_type).font = bold_font
            # Extract and write reference
            ref_without_asterisk = [col for col in snvmatrix_df.columns if col.endswith('*')][0][:-1]
            if rename_mapping is not None and ref_without_asterisk in rename_mapping:
                ref_without_asterisk = rename_mapping[ref_without_asterisk]
            sheet.cell(row=start_row + 2, column=1, value="Reference:")
            sheet.cell(row=start_row + 2, column=2, value=ref_without_asterisk)
            # Write Window Size
            sheet.cell(row=start_row + 3, column=1, value="Window size:")
            sheet.cell(row=start_row + 3, column=2, value=str(window_size))
            # Write core genome percentage
            sheet.cell(row=start_row + 4, column=1, value="SNVPhyl core estimate:")
            sheet.cell(row=start_row + 4, column=2, value=str(result_dict.get(seq_type)) + "%")
            # Write core genome percentage
            sheet.cell(row=start_row + 5, column=1, value="hqSNV Range:")
            sheet.cell(row=start_row + 5, column=2, value=str(snv_range.get(seq_type)))
            # Write header
            for col_idx, column_name in enumerate(snvmatrix_df.columns, start=1):
                if str(column_name).startswith("Unnamed:"):
                    column_name = ""
                sheet.cell(row=start_row + 7, column=col_idx, value=column_name).font = bold_font
            # Write data
            for i, row in snvmatrix_df.iterrows():
                for j, value in enumerate(row):
                    sheet.cell(row=start_row + i + 8, column=j + 1, value=value)
            # Update start_row for next file
            start_row += len(snvmatrix_df) + 8

def get_sorted_files(pattern):
    # Retrieve all matching files
    files = glob.glob(pattern)
    # Separate files containing the All_*_Isolates pattern and those that don't
    All_Isolates_files = [f for f in files if re.search(r'All.*Isolates', f)]
    other_files = [f for f in files if not re.search(r'All.*Isolates', f)]
    # Define function to extract numbers for sorting
    def extract_number(filename):
        match = re.search(r'\d+', filename)
        return int(match.group()) if match else float('inf')  # float('inf') puts files without numbers at the end
    # Sort remaining files by extracted number, then alphabetically for files without numbers
    other_files.sort(key=lambda f: (extract_number(f), f))
    # Concatenate lists with All_Isolates_files first
    return All_Isolates_files + other_files

def get_files():
    # You only need this for glob because glob will throw an index error if not.
    snvmatrices = get_sorted_files("*_snvMatrix.tsv")
    vcf2cores = get_sorted_files("*_vcf2core.tsv")
    #create empty dictionary to fill
    result_dict = {}
    # create snv_range mapping: seq_type -> "min - max"
    snv_range = {}
    # looping through vcf2core files
    for vcf2core in vcf2cores:
        # Derive seq_type from the filename by removing '_vcf2core.tsv'
        seq_type = os.path.basename(vcf2core).replace('_vcf2core.tsv', '')
        # Read the TSV file
        df_vcf2core = pd.read_csv(vcf2core, sep='\t',dtype='str')
        # Extract the last row's specified column value as a float - to get % core genome
        try:
            last_value = float(df_vcf2core["Percentage of all positions that are valid, included, and part of the core genome"].iloc[-1])
            result_dict[seq_type] = last_value
        except (KeyError, ValueError, IndexError) as e:
            print(f"Error processing file '{vcf2core}': {e}")
    # collect SNV min-max ranges
    for snvmatrix in snvmatrices:
        seq_type = os.path.basename(snvmatrix).replace('_snvMatrix.tsv', '')
        try:
            # Read first column as row names
            df = pd.read_csv(snvmatrix, sep='\t', index_col=0, dtype='str')
            # Convert everything to numeric where possible
            df = df.apply(pd.to_numeric, errors='coerce')
            # Make sure row names and column names are comparable strings
            df.index = df.index.astype(str).str.strip()
            df.columns = df.columns.astype(str).str.strip()
            # Keep only shared sample names so matrix is square
            shared_names = [name for name in df.index if name in df.columns]
            df = df.loc[shared_names, shared_names]
            if df.empty or df.shape[0] < 2:
                snv_range[seq_type] = "NA"
                continue
            arr = df.to_numpy(dtype=float)
            # Exclude diagonal and NaN
            non_diag_mask = ~np.eye(arr.shape[0], dtype=bool)
            values = arr[non_diag_mask]
            values = values[~np.isnan(values)]
            if values.size == 0:
                snv_range[seq_type] = "NA"
            else:
                min_val = values.min()
                max_val = values.max()
                # format nicely
                if float(min_val).is_integer():
                    min_val = int(min_val)
                if float(max_val).is_integer():
                    max_val = int(max_val)
                snv_range[seq_type] = f"{min_val} - {max_val}"
                print(f"SNV range for {seq_type}: {snv_range[seq_type]}")
        except Exception as e:
            print(f"Error reading snvMatrix '{snvmatrix}': {e}")
            snv_range[seq_type] = "NA"
    return snvmatrices, result_dict, snv_range

def main():
    args = parseArgs()
    old_griphin = args.griphin
    # Load workbook
    workbook = openpyxl.load_workbook(old_griphin)
    # Create taxa sheets
    taxa_sheet_mapping = create_taxa_sheets(workbook, args.combine_complex)
    # Get SNVMatrix files and core genome data
    snvmatrices, result_dict, snv_range = get_files()
    print(snvmatrices, result_dict)
    # Append SNVPhyl data to taxa sheets
    append_tsv_to_excel(workbook, snvmatrices, result_dict, args.blind_list, taxa_sheet_mapping, args.window_size, snv_range, args.combine_complex)
    # Save the final output file
    workbook.save("SNVPhyl_GRiPHin_Summary.xlsx")
    print("Excel file with taxa sheets and SNVPhyl data saved as 'SNVPhyl_GRiPHin_Summary.xlsx'.")

if __name__ == "__main__":
    main()